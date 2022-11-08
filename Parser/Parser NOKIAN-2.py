import os
import smtplib
import sys
from datetime import datetime
import shutil
import time
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import psycopg2
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.options import Options

from selenium.webdriver.common.by import By

#---------------------подключение к базе и импорт в датафрейм------------------
params_dic = {
	"database": 'Nokian',
	"user": 'postgres',
	"password": '111',
	'port': '5432',
	"host": '127.1.0.0'
}


def connect(connection_data):
	# """ Connect to the PostgreSQL database server """
	conn = None
	try:
		# connect to the PostgreSQL server
		print('Connecting to the PostgreSQL database...')
		conn = psycopg2.connect(**connection_data)
	
	except (Exception, psycopg2.DatabaseError) as error:
		print(error)
		sys.exit(1)
	print("Connection successful")
	return conn


def postgresql_to_dataframe(conn, select_query, column_names):
	# """    # Tranform a SELECT query into a pandas dataframe    # """
	cursor = conn.cursor()
	try:
		cursor.execute(select_query)
	except (Exception, psycopg2.DatabaseError) as error:
		print("Error: %s" % error)
		cursor.close()
		return 1
	# Naturally we get a list of tupples
	tupples = cursor.fetchall()
	cursor.close()
	# We just need to turn it into a pandas dataframe
	df = pd.DataFrame(tupples, columns=column_names)
	return df


conn = connect(params_dic)


#----------------------МОДУЛЬ СКАЧИВАНИЯ ФАЙЛА ОСТАТКОВ С САЙТА НОКИАН----------------------------------


chrome_options = Options()

driver = webdriver.Chrome(r"C:\\Users\\user\\PycharmProjects\\Coursera\\Parser\\chromedriver.exe")
#driver.add_argument('headless')
driver.get('https://egate.nokiantyres.com/wholesale-ru/')

time.sleep(3)
email = driver.find_element(By.ID, "floatingLabelInput19")

email.send_keys('a.dragunov@vistauto.ru')

time.sleep(2)
passwd = driver.find_element(By.ID, 'floatingLabelInput25')
passwd.send_keys('WL6k342F8')

time.sleep(2)
driver.find_element(By.CSS_SELECTOR, ".btn-primary").click()

driver.get('https://egate.nokiantyres.com/wholesale-ru/ru/my-account/warehouses')

try:
	driver.get("https://egate.nokiantyres.com/wholesale-ru/ru/my-account/report.xls")
except:
	raise 'NOTHING'

# C:\Users\user\Downloads
current_time_temp = time.localtime()
current_time = f'{current_time_temp.tm_hour}_{current_time_temp.tm_min}_{current_time_temp.tm_sec}'
current_datetime = datetime.now()
current_date = f'{current_datetime.day}.{current_datetime.month}.{current_datetime.year}'
stock_nokian_filename = f'Stock_NOKIAN {current_date}-{current_time}.xls'



time.sleep(30)
shutil.copyfile(rf'C:\Users\user\Downloads\report.xls', rf'C:\Users\user\PycharmProjects\Coursera\Stock\Price_to_email\{stock_nokian_filename}')

	#print('ФАЙЛ НЕ СКОПИРОВАН ПО КАКИМ-ТО ПРИЧИНАМ')

os.remove('C:\\Users\\user\\Downloads\\report.xls')
driver.quit()
print("Mission complete!")

#---------------конвертирование из xls в xlsx--------

import pandas as pd
df = pd.read_excel(rf'C:\Users\user\PycharmProjects\Coursera\Stock\Price_to_email\{stock_nokian_filename}')
df.to_excel(rf'C:\Users\user\PycharmProjects\Coursera\Stock\Price_to_email\{stock_nokian_filename}x')
os.remove(rf'C:\Users\user\PycharmProjects\Coursera\Stock\Price_to_email\{stock_nokian_filename}')

#---------ИМОРТ ФАЙЛА В БАЗУ-----------------


current_datetime = datetime.now()
database = psycopg2.connect(database="Nokian", user="postgres", password="111", host="localhost", port="5432")
#-----------------------УДАЛЕНИЕ ПРЕДЫДУЩЕГО ПРАЙСА-----------------------------
cursor = database.cursor()
q_delete = f"""delete from nokian_dayly WHERE CUR_MONTH = {current_datetime.month} AND CUR_DAY = {current_datetime.day};"""
cursor.execute(q_delete)
cursor.close()
database.commit()


#-----------------------------------------------------------------------------------------
values = ()
workbook = load_workbook(rf'C:\Users\user\PycharmProjects\Coursera\Stock\Price_to_email\{stock_nokian_filename}x', read_only=True)
worksheet = workbook['Sheet1']

cursor = database.cursor()
query = """INSERT INTO nokian_dayly (
            partnumber,
            brand,
            standard_size,
            tyre_model,
            diameter,
            width,
            height,
            seasons,
            studded,
            runflat,
            name_contract,
            price,
            price_prepaid,
            quantity_msk_sod,
            quantity_pet_svd,
            cur_year,
            cur_month,
            cur_day) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
for row in worksheet.iter_rows(min_row=3, min_col=2, max_col=16, max_row=None, values_only=True):
	print(row)
	values = values + row
	file_month = current_datetime.month  # МЕСЯЦ ЗА КОТОРЫЙ ЗАГРУЖАЕТСЯ ФАЙЛ
	file_day = current_datetime.day  # ДЕНЬ ЗА КОТОРЫЙ ЗАГРУЖАЕТСЯ ФАЙЛ
	# cur_day = 85  # current_datetime.day
	# current_datetime.month
	values = values + (current_datetime.year, file_month, file_day)
	# можно здесь вместо дат вставить обычные числа
	cursor.execute(query, values)
	values = ()

cursor.close()

database.commit()

#-----------выборка из базы ------------------------


cursor = database.cursor()
# query = f"""select partnumber, '', quantity_msk_sod
# 			from nokian_dayly where cur_month = {current_datetime.month} and cur_day= {current_datetime.day} and quantity_msk_sod > 0 and not tyre_model ilike '%nordman%'
# 			and diameter = '15'
# 			union
# 			select partnumber,  '', quantity_msk_sod
# 			from nokian_dayly where cur_month = {current_datetime.month} and cur_day= {current_datetime.day} and quantity_msk_sod > 0 and diameter > '15';"""

query = f"""select partnumber, '', quantity_msk_sod
			from nokian_dayly where cur_month = {current_datetime.month} and cur_day= {current_datetime.day} and quantity_msk_sod > 0;"""



yandex_nokian = f'marketplace-stock NOKIAN {current_date}-{current_time}.xlsx'
column_names = ['partnumber', 'free', 'quantity_msk_sod']
df_nokian = postgresql_to_dataframe(conn, query, column_names)

try:
	shutil.copyfile(rf'C:\Users\user\PycharmProjects\Coursera\Stock\Price_to_email\marketplace-stock YAM.xlsx', rf'C:\Users\user\PycharmProjects\Coursera\Stock\Price_to_email\{yandex_nokian}')
except:
	print('ФАЙЛ НЕ СКОПИРОВАН ПО КАКИМ-ТО ПРИЧИНАМ')

with pd.ExcelWriter(rf'C:\Users\user\PycharmProjects\Coursera\Stock\Price_to_email\{yandex_nokian}', mode='a', if_sheet_exists='overlay') as writer:
	df_nokian.to_excel(writer, sheet_name='Остатки', header=False, index=False, startrow=2, startcol=1)
	
try:
	shutil.copyfile(rf'C:\Users\user\PycharmProjects\Coursera\Stock\Price_to_email\{yandex_nokian}', rf'\\192.168.10.117\kia\ПРАЙС\{yandex_nokian}')
except:
	print('недоступна конечная директория либо проблема с файлом по ШИНАМ СКЛАДА НОКИАН!')
	
#---------------------------отправка письма с просьбой обновить остатки на яндексе------------------------


final_msg = f"Пришло время обновить остаки НОКИАН в ЯНДЕКС - файл {yandex_nokian}"
mail_for_who = "m.lobanov@vistauto.ru"


def send_email(addr_to, msg_subj, msg_text):
    addr_from = "a.dragunov@vistauto.ru"  # Отправитель
    password = "12345678qaZ"  # Пароль

    msg = MIMEMultipart()  # Создаем сообщение
    msg['From'] = "a.dragunov@vistauto.ru"  # Адресат
    msg['To'] = addr_to  # Получатель

    msg['Subject'] = msg_subj  # Тема сообщения

    body = msg_text  # Текст сообщения
    msg.attach(MIMEText(body, 'plain'))  # Добавляем в сообщение текст

    # ======== Этот блок настраивается для каждого почтового провайдера отдельно ===============================================
    server = smtplib.SMTP_SSL('smtp.yandex.ru', 465)  # Создаем объект SMTP
    # server.starttls()                                      # Начинаем шифрованный обмен по TLS
    # server.set_debuglevel(True)                            # Включаем режим отладки, если не нужен - можно закомментировать
    server.login(addr_from, password)  # Получаем доступ
    server.send_message(msg)  # Отправляем сообщение
    server.quit()  # Выходим


#send_email(mail_for_who, final_msg, "")