from openpyxl import load_workbook

values = []

workbook = load_workbook("toyota.xlsx", read_only=True)
worksheet = workbook['TDSheet']

for row in worksheet.iter_rows(min_row=1, max_col=1, max_row=None, values_only=True):
    values.append(row[0])

print(values)
