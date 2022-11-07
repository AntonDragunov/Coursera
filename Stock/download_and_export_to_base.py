import base64
import imaplib
import os
import smtplib
import datetime
import email

import time

import pandas as pd
import psycopg2
import schedule

from openpyxl import load_workbook
import shutil
from zipfile import ZipFile
#from Stock.SEND_MAIL_2 import send_country_list

HOST = "smtp.yandex.ru"
TO = "a.dragunov@vistauto.ru"
FROM = "a.dragunov@vistauto.ru"
password = "12345678qaZ"
text = ""
server = smtplib.SMTP_SSL(HOST, 465)


def mail_price_already_done_here(SUBJECT):
	BODY = "\r\n".join((
		"From: %s" % FROM,
		"To: %s" % TO,
		"Subject: %s" % SUBJECT,
		"",
		text
	))
	server.login(FROM, password)
	server.sendmail(FROM, TO, BODY.format(TO, FROM, SUBJECT, text).encode('utf-8'))
	server.quit()

def there_is_no_price_today_yet(SUBJECT):
	text = "ХАХАХАХАХАХХАХАХАХАХ"
	
	BODY = "\r\n".join((
		"From: %s" % FROM,
		"To: %s" % TO,
		"Subject: %s" % SUBJECT,
		"",
		text
	))
	
	
	server.login(FROM, password)
	server.sendmail(FROM, TO, BODY.format(TO, FROM, SUBJECT, text).encode('utf-8'))
	server.quit()

def download_complete(SUBJECT):

	text = SUBJECT
	
	BODY = "\r\n".join((
		"From: %s" % FROM,
		"To: %s" % TO,
		"Subject: %s" % SUBJECT,
		"",
		text
	))
	
	
	server.login(FROM, password)
	server.sendmail(FROM, TO, BODY.format(TO, FROM, SUBJECT, text).encode('utf-8'))
	server.quit()

def pack_repack():
	# функция необходима для переформатирования файла в редактируемый.
	#вносит незнаичтельные изменения в файл и пересохраняет.
	
	# Создаем временную папку
	tmp_folder = 'convert_wrong_excel'
	os.makedirs(tmp_folder, exist_ok=True)
	
	# Распаковываем excel как zip в нашу временную папку
	with ZipFile('DEALER_PRICE_LIST.xlsx') as excel_container:
		excel_container.extractall(tmp_folder)
	
	# пробуем удалить XLSX
	path = os.path.join(os.path.abspath(os.path.dirname(__file__)), 'DEALER_PRICE_LIST.xlsx')
	os.remove(path)
	
	# Переименовываем файл с неверным названием
	wrong_file_path = os.path.join(tmp_folder, 'xl', 'SharedStrings.xml')
	correct_file_path = os.path.join(tmp_folder, 'xl', 'sharedStrings.xml')
	os.rename(wrong_file_path, correct_file_path)
	
	# Запаковываем excel обратно в zip и переименовываем в исходный файл
	shutil.make_archive('DEALER_PRICE_LIST', 'zip', tmp_folder)
	os.rename('DEALER_PRICE_LIST.zip', 'DEALER_PRICE_LIST.xlsx')

def day():
	mail = imaplib.IMAP4_SSL('imap.yandex.ru')
	mail.login('a.dragunov@vistauto.ru', '12345678qaZ')
	
	mail.list()
	mail.select("KMR")
	detach_dir = ''
	import datetime as datetime
	date = datetime.date.today().strftime("%d-%b-%Y")  # + datetime.timedelta(1))
	result, data = mail.search('None', 'SINCE {date}'.format(date=date))
	
	database = psycopg2.connect(database="Nokian", user="postgres", password="111", host="localhost",
	                            port="5432")
	from datetime import datetime
	current_datetime = datetime.now()
	cursor = database.cursor()
	
	column_names = ['stock']
	IsItpriceHere = f"""SELECT stock FROM kmr_dayly_2 WHERE kmr_dayly_2.cur_day = {current_datetime.day} And kmr_dayly_2.cur_month = {current_datetime.month};"""
	
	cursor.execute(IsItpriceHere)
	tupples = cursor.fetchall()
	df = pd.DataFrame(tupples, columns=column_names)
	# isempty = 0 in df["SUM(stock)"]
	# print('Is the DataFrame empty :', isempty)
	cursor.close()
	database.commit()
	database.close()
	print(len(df))
	
	if len(df) == 1 or len(df) == 0:
		ids = data[0]
		id_list = ids.split()
		print(id_list)
		if id_list:
			# i = 0
			for i in range(len(id_list)):
				
				latest_email_id = id_list[i]
				result, data = mail.fetch(latest_email_id, "(RFC822)")
				raw_email = data[0][1]
				raw_email_string = raw_email.decode('utf-8')
				email_message = email.message_from_string(raw_email_string)
				raw_email_string = raw_email.decode('utf-8')
				from email.header import decode_header
				
				a = []
				for part in decode_header(email_message['Subject']):
					a.append(str(*part))
				print(a)
				if 'Kia Price-list' in a[0]:
					for part in email_message.walk():
						# multipart are just containers, so we skip them
						if part.get_content_maintype() == 'multipart':
							continue
						# is this part an attachment ?
						if part.get('Content-Disposition') is None:
							continue
						filename = part.get_filename()
						counter = 1
						if not filename:
							filename = 'part-%03d%s' % (counter, 'bin')
							counter += 1
						if filename[:11] == '=?KOI8-R?B?':
							filename = base64.b64decode(filename[11:]).decode('KOI8-R')
						if filename[:10] == '=?utf-8?B?':
							filename = base64.b64decode(filename[10:])
						if filename == 'DEALER_PRICE_LIST_DELTA.xlsx':
							print('DELTA')
							continue
							
						else:
							print('PRISE')
							break
					
					filename_real = str(filename)[2:-1]
					att_path = os.path.join(detach_dir, filename_real)
					with open(os.path.join(detach_dir, filename_real), 'wb') as fp:
						fp.write(part.get_payload(decode=True))
					print('Прайс сохранен за' + email_message['Date'])
					# модуль загрузка в базу
					values = ()
					
					# ---------------------------------------------------------------------
					pack_repack()
					# ___________________________________________________________________________
					
					workbook = load_workbook("DEALER_PRICE_LIST.xlsx")
					worksheet = workbook['TDSheet']
					from datetime import datetime
					current_datetime = datetime.now()
					database = psycopg2.connect(database="Nokian", user="postgres", password="111", host="localhost",
					                            port="5432")
					
					cursor = database.cursor()
					query = """INSERT INTO kmr_dayly_2 (
                                    PART_NO,
                                    ITC_CODE,
                                    ITC_PART,
                                    MODEL,
                                    PART_NAME_ENG,
                                    PART_NAME_RUS,
                                    MPQ1,
                                    D_ORDER_DNP,
                                    LIST_PRICE,
                                    STOCK_MC1,
                                    STOCK_ME1,
                                    STOCK_MN1,
                                    STOCK_MS1,
                                    STOCK_MK1,
                                    STOCK,
                                    cur_year,
                                    cur_month,
                                    cur_day) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
					for row in worksheet.iter_rows(min_row=2, max_col=15, max_row=None, values_only=True):
						# print(row)
						values = values + row
						
						# cur_day = 85  # current_datetime.day
						# current_datetime.month
						values = values + (current_datetime.year, current_datetime.month, current_datetime.day)
						# можно здесь вместо дат вставить обычные числа
						cursor.execute(query, values)
						values = ()
					
					cursor.close()
					
					database.commit()
					cursor = database.cursor()
					query2 = """UPDATE kmr_dayly_2 SET part_no = RTRIM(part_no), part_name_rus = RTRIM(part_name_rus);"""
					cursor.execute(query2)
					cursor.close()
					database.commit()
					database.close()
					
					print("I just imported Excel into postgreSQL")
					
					# -------------ОТПРАВКА ПИСЬМА------------------------------------
					download_complete("DOWNLOAD COMPLETE!!!!")
					
					# ----------------------------------------------------------------------------------
					# -----------ЗАПУСК МОДУЛЯ SEND_MAIL-2-------- отправка отчет по дефицитным товарам----------------------
					# send_country_list()
					break
				else:
					print('НЕТ НУЖНОГО ПИСЬМА!')  # pass
				i += 1
		
		
		else:
			print("THERE IS NO EMAIL TODAY!!!!")
			there_is_no_price_today_yet("THERE IS NO EMAIL TODAY!!!!!!!!")
	else:
		print("Price is Here Already , WELL DONE!")
		mail_price_already_done_here("Price is Here Already , WELL DONE!")


schedule.every(1).minutes.do(day)
while True:
	schedule.run_pending()
	time.sleep(1)
