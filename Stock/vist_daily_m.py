import os
import time

import psycopg2

from datetime import datetime
from openpyxl import load_workbook


current_datetime = datetime.now()
file_update_time = time.localtime(os.path.getmtime('free_stock_m.xlsx'))

if file_update_time[1] == current_datetime.month and file_update_time[2] == current_datetime.day:
    values = ()

    workbook = load_workbook("free_stock_m.xlsx", read_only=True)
    worksheet = workbook['TDSheet']

    current_datetime = datetime.now()
    database = psycopg2.connect(database="Nokian", user="postgres", password="111", host="localhost", port="5432")

    cursor = database.cursor()
    query_truncate = """TRUNCATE TABLE vist_daily_m"""
    cursor.execute(query_truncate)
    cursor.close()
    database.commit()

    cursor = database.cursor()
    query = """INSERT INTO vist_daily_m (
                PART_NO,
                free_stock)
                 VALUES (%s, %s)"""
    for row in worksheet.iter_rows(min_row=2, max_col=2, max_row=None, values_only=True):
        print(row)
        values = values + row

        # cur_day = 85  # current_datetime.day
        # current_datetime.month
        # values = values + (current_datetime.year, 4, 25)
        # можно здесь вместо дат вставить обычные числа
        # print(values)
        cursor.execute(query, values)
        values = ()

    cursor.close()

    database.commit()

    database.close()

    print(f'Файл vist_daily_m обновлен - дата:{file_update_time[1]}-{file_update_time[2]}, время:{file_update_time[3]}-{file_update_time[4]},')
    print("I just download vist_daily stocks")


else:
    print('НЕОБХОДИМО ОБНОВИТЬ ФАЙЛ!!!')
