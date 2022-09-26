import os
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import psycopg2

from datetime import datetime
from openpyxl import load_workbook

values = ()

# ______________блок для открытия файла из 1с для записи несущественной информации и его сохранении,
# для того чтобы был возможен импорт без ручного открытия/сохранения----------------------------

import shutil
from zipfile import ZipFile

# Создаем временную папку
tmp_folder = 'convert_wrong_excel'
os.makedirs(tmp_folder, exist_ok=True)

# Распаковываем excel как zip в нашу временную папку
with ZipFile('DEALER_PRICE_LIST.xlsx') as excel_container:
    excel_container.extractall(tmp_folder)

# пробуем удалить XLSX
path = os.path.join(os.path.abspath(os.path.dirname(__file__)), 'DEALER_PRICE_LIST.xlsx')
os.remove(path)

# Переименовываем файл с неверным названием
wrong_file_path = os.path.join(tmp_folder, 'xl', 'SharedStrings.xml')
correct_file_path = os.path.join(tmp_folder, 'xl', 'sharedStrings.xml')
os.rename(wrong_file_path, correct_file_path)

# Запаковываем excel обратно в zip и переименовываем в исходный файл
shutil.make_archive('DEALER_PRICE_LIST', 'zip', tmp_folder)
os.rename('DEALER_PRICE_LIST.zip', 'DEALER_PRICE_LIST.xlsx')

# __________________________________________________________________________

workbook = load_workbook("DEALER_PRICE_LIST.xlsx", read_only=True)
worksheet = workbook['TDSheet']

current_datetime = datetime.now()
database = psycopg2.connect(database="Nokian", user="postgres", password="111", host="localhost", port="5432")

cursor = database.cursor()
query = """INSERT INTO kmr_dayly (
            PART_NO,
            ITC_CODE,
            ITC_PART,
            MODEL,
            PART_NAME_ENG,
            PART_NAME_RUS,
            MPQ1,
            D_ORDER_DNP,
            LIST_PRICE,
            STOCK_MC1,
            STOCK_ME1,
            STOCK_MN1,
            STOCK_MS1,
            STOCK_MK1,
            STOCK,
            cur_year,
            cur_month,
            cur_day) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
for row in worksheet.iter_rows(min_row=2, max_col=15, max_row=None, values_only=True):
    print(row)
    values = values + row
    file_month = 9  # МЕСЯЦ ЗА КОТОРЫЙ ЗАГРУЖАЕТСЯ ФАЙЛ
    file_day = 26  # ДЕНЬ ЗА КОТОРЫЙ ЗАГРУЖАЕТСЯ ФАЙЛ
    # cur_day = 85  # current_datetime.day
    # current_datetime.month
    values = values + (current_datetime.year, file_month, file_day)
    # можно здесь вместо дат вставить обычные числа
    cursor.execute(query, values)
    values = ()

cursor.close()

database.commit()
# -------------удаление лишних пробелов в номере детали-----------------------


cursor = database.cursor()
query2 = f"""UPDATE kmr_dayly SET part_no = RTRIM(part_no) WHERE CUR_MONTH = {file_month} AND CUR_DAY = {file_day}"""
cursor.execute(query2)
cursor.close()
database.commit()

# ----------------------------------------------------------------------------
final_msg = f"СТОК КМР ИМПОРТИРОВАН В БАЗУ - за {file_day}.{file_month}.{current_datetime.year}"
database.close()
print()
print()
print(final_msg)


def send_email(addr_to, msg_subj, msg_text):
    addr_from = "a.dragunov@vistauto.ru"  # Отправитель
    password = "12345678qaZ"  # Пароль

    msg = MIMEMultipart()  # Создаем сообщение
    msg['From'] = "a.dragunov@vistauto.ru"  # Адресат
    msg['To'] = addr_to  # Получатель

    msg['Subject'] = msg_subj  # Тема сообщения

    body = msg_text  # Текст сообщения
    msg.attach(MIMEText(body, 'plain'))  # Добавляем в сообщение текст

    # ======== Этот блок настраивается для каждого почтового провайдера отдельно ===============================================
    server = smtplib.SMTP_SSL('smtp.yandex.ru', 465)  # Создаем объект SMTP
    # server.starttls()                                      # Начинаем шифрованный обмен по TLS
    # server.set_debuglevel(True)                            # Включаем режим отладки, если не нужен - можно закомментировать
    server.login(addr_from, password)  # Получаем доступ
    server.send_message(msg)  # Отправляем сообщение
    server.quit()  # Выходим


send_email("a.dragunov@vistauto.ru", final_msg, "")
