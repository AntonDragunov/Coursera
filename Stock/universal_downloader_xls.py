import psycopg2

from datetime import datetime
from openpyxl import load_workbook

values = ()

workbook = load_workbook("tyre_brands.xlsx", read_only=True)
worksheet = workbook['Лист1']

current_datetime = datetime.now()
database = psycopg2.connect(database="Nokian", user="postgres", password="111", host="localhost", port="5432")

cursor = database.cursor()
query_truncate = """TRUNCATE TABLE tyre_brands"""
cursor.execute(query_truncate)
cursor.close()
database.commit()

cursor = database.cursor()
query = """INSERT INTO tyre_brands (
            BRAND)
             VALUES (%s)"""
for row in worksheet.iter_rows(min_row=1, max_col=1, max_row=None, values_only=True):
    print(row)
    values = values + row

    # cur_day = 85  # current_datetime.day
    # current_datetime.month
    # values = values + (current_datetime.year, 4, 25)
    # можно здесь вместо дат вставить обычные числа
    # print(values)
    cursor.execute(query, values)
    values = ()

cursor.close()

database.commit()

database.close()

print("I just imported Excel into postgreSQL")
