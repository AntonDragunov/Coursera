import base64
import email
import imaplib
import mimetypes
import os
import smtplib
from email import encoders
from email.header import Header
from email.mime.audio import MIMEAudio
from email.mime.base import MIMEBase
from email.mime.image import MIMEImage
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import pandas as pd
import psycopg2
from openpyxl import load_workbook



def send_email_4tochki(addr_to, msg_subj, msg_text, files):
    addr_from = "a.dragunov@vistauto.ru"  # Отправитель
    password = "12345678qaZ"  # Пароль

    msg = MIMEMultipart()  # Создаем сообщение
    msg['From'] = "a.dragunov@vistauto.ru"  # Адресат
    msg['To'] = "a.dragunov@vistauto.ru"  # Получатель
    msg['Cc'] = "a.dragunov@vistauto.ru"  # COPY
    msg['Subject'] = msg_subj  # Тема сообщения

    body = msg_text  # Текст сообщения
    msg.attach(MIMEText(body, 'plain'))  # Добавляем в сообщение текст

    process_attachement(msg, files)

    # ======== Этот блок настраивается для каждого почтового провайдера отдельно ===============================================
    server = smtplib.SMTP_SSL('smtp.yandex.ru', 465)  # Создаем объект SMTP
    # server.starttls()                                      # Начинаем шифрованный обмен по TLS
    # server.set_debuglevel(True)                            # Включаем режим отладки, если не нужен - можно закомментировать
    server.login(addr_from, password)  # Получаем доступ
    server.send_message(msg)  # Отправляем сообщение
    server.quit()  # Выходим
    # ==========================================================================================================================


def process_attachement(msg, files):  # Функция по обработке списка, добавляемых к сообщению файлов
    for f in files:
        if os.path.isfile(f):  # Если файл существует
            attach_file(msg, f)  # Добавляем файл к сообщению
        elif os.path.exists(f):  # Если путь не файл и существует, значит - папка
            dir = os.listdir(f)  # Получаем список файлов в папке
            for file in dir:  # Перебираем все файлы и...
                attach_file(msg, f + "/" + file)  # ...добавляем каждый файл к сообщению


def attach_file(msg, filepath):  # Функция по добавлению конкретного файла к сообщению
    filename = os.path.basename(filepath)  # Получаем только имя файла
    ctype, encoding = mimetypes.guess_type(filepath)  # Определяем тип файла на основе его расширения
    if ctype is None or encoding is not None:  # Если тип файла не определяется
        ctype = 'application/octet-stream'  # Будем использовать общий тип
    maintype, subtype = ctype.split('/', 1)  # Получаем тип и подтип
    if maintype == 'text':  # Если текстовый файл
        with open(filepath) as fp:  # Открываем файл для чтения
            file = MIMEText(fp.read(), _subtype=subtype)  # Используем тип MIMEText
            fp.close()  # После использования файл обязательно нужно закрыть
    elif maintype == 'image':  # Если изображение
        with open(filepath, 'rb') as fp:
            file = MIMEImage(fp.read(), _subtype=subtype)
            fp.close()
    elif maintype == 'audio':  # Если аудио
        with open(filepath, 'rb') as fp:
            file = MIMEAudio(fp.read(), _subtype=subtype)
            fp.close()
    else:  # Неизвестный тип файла
        with open(filepath, 'rb') as fp:
            file = MIMEBase(maintype, subtype)  # Используем общий MIME-тип
            file.set_payload(fp.read())  # Добавляем содержимое общего типа (полезную нагрузку)
            fp.close()
            encoders.encode_base64(file)  # Содержимое должно кодироваться как Base64

    mail_coding = 'utf-8'  # Добавляем заголовки
    att_header = Header(os.path.basename(filepath), mail_coding);
    file.add_header('Content-Disposition', 'attachment; filename="%s"' % att_header)

    msg.attach(file)  # Присоединяем файл к сообщению






def day():
    mail = imaplib.IMAP4_SSL('imap.yandex.ru')
    mail.login('a.dragunov@vistauto.ru', '12345678qaZ')

    mail.list()
    mail.select('INBOX')
    detach_dir = ''
    import datetime as datetime
    date = datetime.date.today().strftime("%d-%b-%Y")  # + datetime.timedelta(1))
    result, data = mail.search('None', 'SINCE {date}'.format(date=date))


    # database = psycopg2.connect(database="Nokian", user="postgres", password="111", host="localhost",
    #                                         port="5432")
    from datetime import datetime
    current_datetime = datetime.now()
    # cursor = database.cursor()

    # column_names = ['stock']
    # IsItpriceHere = f"""SELECT stock FROM kmr_dayly_2 WHERE kmr_dayly_2.cur_day = {current_datetime.day} And kmr_dayly_2.cur_month = {current_datetime.month};"""
    today_title = 'Форточки: Автоматическая рассылка остатков'
    # cursor.execute(IsItpriceHere)
    # tupples = cursor.fetchall()
    # df = pd.DataFrame(tupples, columns=column_names)
    # #isempty = 0 in df["SUM(stock)"]
    # #print('Is the DataFrame empty :', isempty)
    # cursor.close()
    # database.commit()
    # database.close()
    # print(len(df))


    ids = data[0]
    id_list = ids.split()
    print(id_list)
    for element in id_list:
        print(type(element))
    if id_list:
        i = 0
        for element in id_list:
            latest_email_id = id_list[i]
            result, data = mail.fetch(latest_email_id, "(RFC822)")
            raw_email = data[0][1]
            raw_email_string = raw_email.decode('utf-8')
            email_message = email.message_from_string(raw_email_string)
            raw_email_string = raw_email.decode('utf-8')
            from email.header import decode_header

            a = []
            for part in decode_header(email_message['Subject']):
                a.append(str(*part))
            print(a[0])

            if today_title in a[0]:
                for part in email_message.walk():
                    # multipart are just containers, so we skip them
                    if part.get_content_maintype() == 'multipart':
                        continue
                        # is this part an attachment ?
                    if part.get('Content-Disposition') is None:
                        continue
                    filename = part.get_filename()
                    counter = 1
                    if not filename:
                        filename = 'part-%03d%s' % (counter, 'bin')
                        counter += 1

                if filename[:11] == '=?KOI8-R?B?':
                    filename = base64.b64decode(filename[11:]).decode('KOI8-R')
                if filename[:10] == '=?utf-8?B?':
                    filename = base64.b64decode(filename[10:])

                filename_real = str(filename)[:]
                att_path = os.path.join(detach_dir, filename_real)
                with open(os.path.join(detach_dir, filename_real), 'wb') as fp:
                    fp.write(part.get_payload(decode=True))
                #print('Прайс сохранен за ' + email_message['Date'])
                    # модуль загрузка в базу
                    # values = ()
                    #
                    # # ---------------------------------------------------------------------
                    # # модуль shutil удалил
                    # # ___________________________________________________________________________
                    #
                    # workbook = load_workbook("filename_real.csv")
                    # worksheet = workbook['filename_real']
                    # from datetime import datetime
                    # current_datetime = datetime.now()
                    # database = psycopg2.connect(database="Nokian", user="postgres", password="111", host="localhost",
                    #                             port="5432")
                    #
                    # cursor = database.cursor()
                    # query = """INSERT INTO kmr_dayly_2 (
                    #                 PART_NO,
                    #                 ITC_CODE,
                    #                 ITC_PART,
                    #                 MODEL,
                    #                 PART_NAME_ENG,
                    #                 PART_NAME_RUS,
                    #                 MPQ1,
                    #                 D_ORDER_DNP,
                    #                 LIST_PRICE,
                    #                 STOCK_MC1,
                    #                 STOCK_ME1,
                    #                 STOCK_MN1,
                    #                 STOCK_MS1,
                    #                 STOCK_MK1,
                    #                 STOCK,
                    #                 cur_year,
                    #                 cur_month,
                    #                 cur_day) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
                    # for row in worksheet.iter_rows(min_row=2, max_col=15, max_row=None, values_only=True):
                    #     # print(row)
                    #     values = values + row
                    #
                    #     # cur_day = 85  # current_datetime.day
                    #     # current_datetime.month
                    #     values = values + (current_datetime.year, current_datetime.month, current_datetime.day)
                    #     # можно здесь вместо дат вставить обычные числа
                    #     cursor.execute(query, values)
                    #     values = ()
                    #
                    # cursor.close()
                    #
                    # database.commit()
                    # cursor = database.cursor()
                    # query2 = """UPDATE kmr_dayly_2 SET part_no = RTRIM(part_no), part_name_rus = RTRIM(part_name_rus);"""
                    # cursor.execute(query2)
                    # cursor.close()
                    # database.commit()
                    # database.close()

                    #print("I just imported Excel into postgreSQL")

                    # -------------ОТПРАВКА ПИСЬМА------------------------------------

                # HOST = "smtp.yandex.ru"
                # SUBJECT = "THIS IS EMAIL"
                # TO = "a.dragunov@vistauto.ru"
                # FROM = "a.dragunov@vistauto.ru"
                # text = "ПРАЙС ПЕРЕСЛАН!"
                #
                # BODY = "\r\n".join((
                #         "From: %s" % FROM,
                #         "To: %s" % TO,
                #         "Subject: %s" % SUBJECT,
                #         "",
                #         text
                # ))
                #
                # server = smtplib.SMTP_SSL(HOST, 465)
                # server.login('a.dragunov@vistauto.ru', '12345678qaZ')
                # server.sendmail(FROM, TO, BODY.format(TO, FROM, SUBJECT, text).encode('utf-8'))
                # server.quit()
                    # ----------------------------------------------------------------------------------
                    #-----------ЗАПУСК МОДУЛЯ SEND_MAIL-2-------- отправка отчет по дефицитным товарам----------------------
                files = [f'{filename}']  # Если нужно отправить все файлы из заданной папки, нужно указать её
                print(f'{filename}', i)
                #send_email_4tochki('a.dragunov@vistauto.ru', '    ',f'{filename}', files)

            else:
                print('THERE IS NOTHING EMAIL!')#pass
            i += 1


    else:
        print("THERE IS NO EMAIL TODAY!!!!")
        HOST = "smtp.yandex.ru"
        SUBJECT = "THERE IS NO EMAIL TODAY!!!!!!!!"
        TO = "a.dragunov@vistauto.ru"
        FROM = "a.dragunov@vistauto.ru"
        text = "ХАХАХАХАХАХХАХАХАХАХ"

        BODY = "\r\n".join((
                "From: %s" % FROM,
                "To: %s" % TO,
                "Subject: %s" % SUBJECT,
                "",
                text
        ))

        server = smtplib.SMTP_SSL(HOST, 465)
        server.login('a.dragunov@vistauto.ru', '12345678qaZ')
        server.sendmail(FROM, TO, BODY.format(TO, FROM, SUBJECT, text).encode('utf-8'))
        server.quit()


day()