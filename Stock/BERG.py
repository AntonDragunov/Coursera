import json

import requests
from openpyxl import Workbook
from openpyxl import load_workbook
list_of_partnambers = []
values = []


### ----------------НУЖНО ВСТАВИТЬ АТКУАЛЬНЫЕ ИМЕНЯ ВХОДНОГО И ВЫХОДНОГО ФАЙЛОВ

input_file = 'PIK_VAZ.xlsx'
output_file = 'PIK_VAZ_result.xlsx'

workbook = load_workbook(f'{input_file}', read_only=True)
worksheet = workbook['TDSheet']

for row in worksheet.iter_rows(min_row=1, max_col=1, max_row=None, values_only=True):
    list_of_partnambers.append(str(row[0]))
    #print(list_of_partnambers)


#https://api.berg.ru/ordering/get_stock.xml?items[0][resource_article]=GDB1044&items[1][resource_article]=251711321&key=f5d161de3c926452391ac5948119470249502f451f79251370a49e56fe6afe2a


# print(values)

# partnumber = 'GDB1044'
# new_access_token = 'f5d161de3c926452391ac5948119470249502f451f79251370a49e56fe6afe2a'
# url = f'https://api.berg.ru/ordering/get_stock.json?items[0][resource_article]=GDB1044&key=f5d161de3c926452391ac5948119470249502f451f79251370a49e56fe6afe2a'
#
# response = requests.get(url)
# data = response.json()
# print((data['resources'][0])['article'])
# print((data['resources'][0])['name'])
# print((data['resources'][0])['brand']['name'])
#
# suppliers_quantytу = 4
# for i in range(suppliers_quantytу):
#     print('поставщик ' + str((i + 1)))
#     print('цена', (data['resources'][0])['offers'][i]['price'])
#     print('количество', (data['resources'][0])['offers'][i]['quantity'])
#     print('срок поставки', (data['resources'][0])['offers'][i]['average_period'])
#     print()
# пробую записать все данные по одному номеру в словарь

#list_of_partnambers = ['04E129620', '5Q0129620B', '1K0129620D', '5Q0819669', '6R0819653', '06L115562B', '04E115561H',
                       #  '06J115403Q', '04E129620A', '06A115561B', '4853006A50', '1K1819653B']
output_data = [['артикул', 'производитель', 'наименование', 'поставщик', 'цена', 'количество', 'срок поставки']]

#list_of_partnambers = values
#print(list_of_partnambers)


def one_number(partnumber, suppliers_quantytу):
    new_access_token = 'f5d161de3c926452391ac5948119470249502f451f79251370a49e56fe6afe2a'
    url = f'https://api.berg.ru/ordering/get_stock.json?items[0][resource_article]={partnumber}&key={new_access_token}'
    response = requests.get(url)
    data = response.json()
    #print(data)
    list1 = []
    if len(data['resources']) != 0:#(data['resources'][0]).get('article', False):
        list1.append((data['resources'][0])['article'])
        list1.append((data['resources'][0])['brand']['name'])
        list1.append((data['resources'][0])['name'])
        for i in range(suppliers_quantytу):
            list1.append((data['resources'][0])['offers'][i]['warehouse']['name'])
            list1.append((data['resources'][0])['offers'][i]['price'])
            list1.append((data['resources'][0])['offers'][i]['quantity'])
            list1.append((data['resources'][0])['offers'][i]['average_period'])
        output_data.append(list1)
    else:
        list1.append(partnumber)
        output_data.append(list1)


for i in list_of_partnambers:
    #print(i)
    one_number(i, 1)

book = Workbook()
sheet = book.active

for _ in output_data:
    sheet.append(tuple(_))
    print(*_, end='\n')

book.save(f'{output_file}')